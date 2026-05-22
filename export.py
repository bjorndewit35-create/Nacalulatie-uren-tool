"""Bouwt het nacalculatie-overzicht als Excel (.xlsx) of CSV in het geheugen."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font

from parsing import min_naar_hhmm, min_naar_uur

KOPPEN = [
    "Medewerker", "Datum", "Functie", "Begintijd", "Eindtijd",
    "Plantijd (uur)", "Werkelijk die dag (uur)", "Toegekend (uur)",
    "Toegekend (uu:mm)", "Bron", "Status", "Opmerking",
]


def _rijen(resultaat):
    for m in resultaat["medewerkers"]:
        for r in m["regels"]:
            yield [
                m["naam"],
                r["datum"],
                r["functie"],
                r["begintijd"] or "",
                r["eindtijd"] or "",
                min_naar_uur(r["plantijd_min"]),
                min_naar_uur(r["werkelijk_dag_min"]) if r["werkelijk_dag_min"] is not None else "",
                min_naar_uur(r["toegekend_min"]),
                min_naar_hhmm(r["toegekend_min"]),
                r["bron"],
                r["status"],
                r["opmerking"],
            ]


def naar_xlsx(resultaat):
    wb = Workbook()
    ws = wb.active
    ws.title = "Nacalculatie"

    titel = resultaat.get("projectnaam") or "Nacalculatie"
    ws.append([f"Nacalculatie: {titel}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(KOPPEN)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    for rij in _rijen(resultaat):
        ws.append(rij)

    ws.append([])
    tot = resultaat["project_totaal_min"]
    ws.append(["PROJECTTOTAAL", "", "", "", "", "", "", min_naar_uur(tot), min_naar_hhmm(tot)])
    ws[ws.max_row][0].font = Font(bold=True)

    if resultaat["ongematcht"]:
        ws.append([])
        ws.append(["Ongematchte planningregels (controleren: spelfout of alias?)"])
        ws[ws.max_row][0].font = Font(bold=True)
        ws.append(["Naam", "Datum", "Functie"])
        for pr in resultaat["ongematcht"]:
            ws.append([pr["werknemer"], pr["datum"], pr["functie"]])

    for col, breedte in zip("ABCDEFGHIJKL", [22, 12, 30, 10, 10, 12, 18, 14, 16, 18, 16, 40]):
        ws.column_dimensions[col].width = breedte

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def naar_csv(resultaat):
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(KOPPEN)
    for rij in _rijen(resultaat):
        w.writerow(rij)
    w.writerow([])
    tot = resultaat["project_totaal_min"]
    w.writerow(["PROJECTTOTAAL", "", "", "", "", "", "", min_naar_uur(tot), min_naar_hhmm(tot)])
    data = buf.getvalue().encode("utf-8-sig")
    return io.BytesIO(data)
